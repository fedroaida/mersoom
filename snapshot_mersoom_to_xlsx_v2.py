#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
snapshot_mersoom_to_xlsx_v3.py (v3-hotfix1)

Purpose
- Reads mersoom_*.json files in a directory (produced by mersoom_agent v18.0+ (incl. v18.0 hotfix))
- Computes an "IQ-like" Growth score + multi-dimensional benchmark
- Appends a snapshot row into an Excel scorecard (xlsx)
- Keeps ONE xlsx file and appends cumulatively (schema-aware append)
- Rebuilds Charts sheet so chart ranges always include latest rows

Compatibility focus
- Designed to work with mersoom_agent v18.x state/policy/semantic/brain schemas (backward-compatible with v16.5+ where possible).

Default directory:
  D:\강준규\업무정리\##개인작업

Run:
  python snapshot_mersoom_to_xlsx_v2_hotfix1.py
or:
  python snapshot_mersoom_to_xlsx_v2_hotfix1.py --dir "D:\강준규\업무정리\##개인작업"

Output:
- Default: <dir>\mersoom_agent_iq_scorecard.xlsx
- Or via env: MERSOOM_SCORECARD
- Or via CLI:  --out <path>

Notes
- The agent's mersoom_memory.json is a rolling window (recent only).
  Lifetime (누적) metrics are derived primarily from:
  - mersoom_state.json: total_actions / evaluated_count / total_reward / voted_posts etc.
  - mersoom_semantic.json: by_day counters (votes/posts/comments/eval/ticks)
  - mersoom_memory_archive.jsonl (if present): lifetime category/style usage for contributions
"""

from __future__ import annotations

import argparse
import datetime
import json
import math
import os
from pathlib import Path
from collections import Counter
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart
from openpyxl.chart.reference import Reference
from openpyxl.chart.radar_chart import RadarChart
from openpyxl.worksheet.worksheet import Worksheet


# ---- Timezone: KST ----
KST = datetime.timezone(datetime.timedelta(hours=9))


def kst_now_naive() -> datetime.datetime:
    """Return current time in KST, timezone removed (Excel-friendly)."""
    return (
        datetime.datetime.utcnow()
        .replace(tzinfo=datetime.timezone.utc)
        .astimezone(KST)
        .replace(tzinfo=None)
    )


# ---- helpers ----
def clamp(x: float, lo: float = 0.0, hi: float = 100.0) -> float:
    return max(lo, min(hi, x))


def shannon_entropy(counter: Counter) -> float:
    total = sum(counter.values())
    if total <= 0:
        return 0.0
    ent = 0.0
    for c in counter.values():
        p = c / total
        ent -= p * math.log(p + 1e-12)
    return ent


def norm_entropy(counter: Counter) -> float:
    n = len(counter)
    if n <= 1:
        return 0.0
    return shannon_entropy(counter) / math.log(n)


def load_json(path: Path, default: Any) -> Any:
    """Best-effort JSON loader (safe under concurrent writes)."""
    try:
        if not path.exists():
            return default
        txt = path.read_text(encoding="utf-8")
        return json.loads(txt)
    except Exception:
        return default


def path_from_env(files_dir: Path, env_key: str, default_name: str) -> Path:
    """Resolve a file path from env var; relative paths are treated as relative to files_dir."""
    v = os.environ.get(env_key)
    if v:
        p = Path(v)
        if not p.is_absolute():
            p = (files_dir / p).resolve()
        return p
    return (files_dir / default_name)


def iter_jsonl(path: Path) -> Iterable[Dict[str, Any]]:
    """Yield dict per JSONL line; silently skips bad lines."""
    if not path.exists():
        return
    try:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict):
                        yield obj
                except Exception:
                    continue
    except Exception:
        return


def sum_semantic_by_day(semantic: Dict[str, Any]) -> Dict[str, float]:
    """
    Aggregates semantic['by_day'][day] counters.
    Expected keys (v16.x):
      ticks, post, comment, eval, vote:up, vote:down
    """
    by_day = semantic.get("by_day") or {}
    totals: Counter = Counter()
    if not isinstance(by_day, dict):
        return dict(totals)

    for _, d in by_day.items():
        if not isinstance(d, dict):
            continue
        for k, v in d.items():
            try:
                totals[k] += float(v)
            except Exception:
                continue
    return dict(totals)


def policy_drift_fraction(policy_obj: Dict[str, Any]) -> float:
    """
    Fraction of policy weights that drifted from 1.0 (signal that learning updates happened).
    Handles v16.x policy shape:
      strategy / comment_length / tone / reply_styles / post_styles (dict of float)
      context overrides: policy['context'][bucket@ctx] = {arm: weight}
      templates: policy['templates']['items'][tid]['weight']
    """
    drift = 0
    total = 0

    def scan_weight_dict(d: Dict[str, Any]) -> None:
        nonlocal drift, total
        for _, w in d.items():
            if isinstance(w, (int, float)):
                total += 1
                if abs(float(w) - 1.0) > 1e-9:
                    drift += 1

    for bucket in ("strategy", "comment_length", "tone", "action_type", "reply_styles", "post_styles"):
        v = policy_obj.get(bucket)
        if isinstance(v, dict):
            scan_weight_dict(v)

    ctx = policy_obj.get("context")
    if isinstance(ctx, dict):
        for _, v in ctx.items():
            if isinstance(v, dict):
                scan_weight_dict(v)

    templates = policy_obj.get("templates")
    if isinstance(templates, dict):
        items = templates.get("items")
        if isinstance(items, dict):
            for _, obj in items.items():
                if isinstance(obj, dict) and isinstance(obj.get("weight"), (int, float)):
                    total += 1
                    if abs(float(obj["weight"]) - 1.0) > 1e-9:
                        drift += 1

    return (drift / total) if total else 0.0


def count_policy_options(policy_obj: Dict[str, Any]) -> int:
    """Counts total number of selectable arms."""
    total = 0
    for bucket in ("strategy", "comment_length", "tone", "action_type", "reply_styles", "post_styles"):
        v = policy_obj.get(bucket)
        if isinstance(v, dict):
            total += len(v)

    templates = policy_obj.get("templates")
    if isinstance(templates, dict):
        items = templates.get("items")
        if isinstance(items, dict):
            total += len(items)

    return total


def style_usage_from_archive(archive_path: Path) -> set[str]:
    """
    Tracks which policy knobs were actually used over lifetime contributions (post/comment)
    using mersoom_memory_archive.jsonl when present.
    """
    used: set[str] = set()
    for it in iter_jsonl(archive_path):
        act = str(it.get("action") or "")
        if act in ("comment", "reply"):
            s = str(it.get("used_strategy") or "")
            t = str(it.get("used_tone") or "")
            l = str(it.get("used_length") or "")
            tid = str(it.get("template_id") or "")
            if s:
                used.add(f"strategy:{s}")
            if t:
                used.add(f"tone:{t}")
            if l:
                used.add(f"comment_length:{l}")
            if tid:
                used.add(f"template:{tid}")
        elif act == "post":
            sty = str(it.get("used_style") or "")
            if sty:
                used.add(f"post_styles:{sty}")
    return used


def category_counter_from_archive(archive_path: Path) -> Counter:
    cats: Counter = Counter()
    for it in iter_jsonl(archive_path):
        if str(it.get("action") or "") not in ("post", "comment", "reply"):
            continue
        cats[str(it.get("category") or "")] += 1
    return cats


def action_counter_lifetime(state: Dict[str, Any], semantic_totals: Dict[str, float]) -> Counter:
    """
    Lifetime action counts (approx).
    Primary source: semantic_totals aggregated from semantic['by_day'] (restart-safe).
    Fallbacks:
      - votes: can be derived from state['voted_posts'] (unique votes) if semantic is empty.
      - if everything is missing, fall back to state['total_actions'] as 'unknown'.
    """
    acts: Counter = Counter()

    votes_up = int(semantic_totals.get("vote:up", 0) or 0)
    votes_down = int(semantic_totals.get("vote:down", 0) or 0)

    posts = int(semantic_totals.get("post", 0) or 0)

    # v18.x can split replies into reply_other / reply_own
    comments_root = int(semantic_totals.get("comment", 0) or 0)
    reply_other = int(semantic_totals.get("reply_other", 0) or 0)
    reply_own = int(semantic_totals.get("reply_own", 0) or 0)

    # older/alternate key (just in case)
    reply_generic = int(semantic_totals.get("reply", 0) or 0)
    if reply_generic and (reply_other == 0 and reply_own == 0):
        reply_other = reply_generic

    # semantic -> acts
    acts["vote:up"] = votes_up
    acts["vote:down"] = votes_down
    acts["post"] = posts
    acts["comment"] = comments_root
    acts["reply_other"] = reply_other
    acts["reply_own"] = reply_own

    # vote fallback: state['voted_posts'] stores per-post vote type
    if (votes_up + votes_down) == 0:
        vp = state.get("voted_posts") if isinstance(state, dict) else None
        if isinstance(vp, dict) and vp:
            cu = 0
            cd = 0
            for _, info in vp.items():
                if not isinstance(info, dict):
                    continue
                t = str(info.get("type") or "")
                if t == "up":
                    cu += 1
                elif t == "down":
                    cd += 1
            acts["vote:up"] = max(acts["vote:up"], cu)
            acts["vote:down"] = max(acts["vote:down"], cd)

    # if semantic is effectively empty, fall back to total_actions
    if sum(acts.values()) == 0:
        total = int((state or {}).get("total_actions", 0) or 0)
        if total:
            acts["unknown"] = total

    return acts



def compute_recent_eval_success(memory_items: List[Dict[str, Any]]) -> Optional[float]:
    """
    Recent-window proxy for evaluation success.
    - Consider evaluated post/comment items in memory.json.
    - Count as 'success' if metrics_after is non-empty OR reward_features exists.
    """
    total = 0
    ok = 0
    for it in memory_items:
        if not isinstance(it, dict):
            continue
        act = str(it.get("action") or "")
        if act not in ("post", "comment", "reply"):
            continue
        if it.get("evaluated") is not True:
            continue
        total += 1
        after = it.get("metrics_after") or {}
        feats = it.get("reward_features")
        if (isinstance(after, dict) and len(after) > 0) or isinstance(feats, dict):
            ok += 1
    return (ok / total) if total else None


def compute_recent_reward_stats(memory_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Recent-window reward stats from memory.json (evaluated post/comment entries)."""
    rs: List[float] = []
    for it in memory_items:
        if not isinstance(it, dict):
            continue
        act = str(it.get("action") or "")
        if act not in ("post", "comment", "reply"):
            continue
        if it.get("evaluated") is not True:
            continue
        r = it.get("reward_scalar")
        if isinstance(r, (int, float)):
            rs.append(float(r))

    if not rs:
        return {"n": 0, "mean": None, "pos_rate": None}

    mean = sum(rs) / len(rs)
    pos_rate = sum(1 for x in rs if x > 0) / len(rs)
    return {"n": len(rs), "mean": mean, "pos_rate": pos_rate}


# ---- snapshot builder ----
def build_snapshot(files_dir: Path) -> Dict[str, Any]:
    brain_path = path_from_env(files_dir, "MERSOOM_BRAIN", "mersoom_brain.json")
    policy_path = path_from_env(files_dir, "MERSOOM_POLICY", "mersoom_policy.json")
    semantic_path = path_from_env(files_dir, "MERSOOM_SEMANTIC", "mersoom_semantic.json")
    state_path = path_from_env(files_dir, "MERSOOM_STATE", "mersoom_state.json")
    memory_path = path_from_env(files_dir, "MERSOOM_MEMORY", "mersoom_memory.json")
    archive_path = path_from_env(files_dir, "MERSOOM_MEMORY_ARCHIVE", "mersoom_memory_archive.jsonl")

    brain = load_json(brain_path, {})
    policy = load_json(policy_path, {})
    semantic = load_json(semantic_path, {})
    state = load_json(state_path, {})
    memory = load_json(memory_path, [])

    if not isinstance(memory, list):
        memory = []

    now = kst_now_naive()
    date_kst = now.date().isoformat()

    # Lifetime counters
    sem_tot = sum_semantic_by_day(semantic if isinstance(semantic, dict) else {})
    acts_life = action_counter_lifetime(state if isinstance(state, dict) else {}, sem_tot)

    total_actions_life = int((state or {}).get("total_actions", 0) or 0)
    if total_actions_life <= 0:
        total_actions_life = int(sum(acts_life.values()))

    votes_life = int(acts_life.get("vote:up", 0) + acts_life.get("vote:down", 0))
    posts_life = int(acts_life.get("post", 0))

    # v18.x: replies can be tracked separately
    comments_root_life = int(acts_life.get("comment", 0))
    replies_other_life = int(acts_life.get("reply_other", 0))
    replies_own_life = int(acts_life.get("reply_own", 0))
    comments_life = comments_root_life + replies_other_life + replies_own_life

    total_contrib_life = posts_life + comments_life

    # Recent-window counters
    acts_recent: Counter = Counter()
    comments_root_recent = 0
    replies_recent = 0
    replies_own_recent = 0
    replies_other_recent = 0

    for it in memory:
        if not isinstance(it, dict):
            continue
        a = str(it.get("action") or "")
        acts_recent[a] += 1

        if a == "comment":
            # root comment vs (legacy) comment-as-reply
            if it.get("parent_id"):
                replies_recent += 1
            else:
                comments_root_recent += 1

        elif a == "reply":
            replies_recent += 1
            at = str(it.get("action_type") or "")
            if at == "reply_own":
                replies_own_recent += 1
            elif at == "reply_other":
                replies_other_recent += 1


    # EXP: experience grows with lifetime actions (log scaled)
    experience = clamp(100 * math.log10(1 + total_actions_life) / 2.5)

    # JDG: lifetime average reward from state
    reward_clip = float(os.environ.get("MERSOOM_REWARD_CLIP", 3.0))
    evaluated_count = int((state or {}).get("evaluated_count", 0) or 0)
    total_reward = float((state or {}).get("total_reward", 0.0) or 0.0)
    avg_reward = (total_reward / evaluated_count) if evaluated_count > 0 else None
    if avg_reward is None:
        judgment = 0.0
    else:
        avg_reward_c = max(-reward_clip, min(reward_clip, float(avg_reward)))
        judgment = clamp(50.0 + 50.0 * (avg_reward_c / reward_clip))

    # DEP: depth of interaction (contributions share vs lifetime)
    engagement = clamp(100 * (total_contrib_life / total_actions_life) / 0.35) if total_actions_life else 0.0

    # DIV: diversity (category/action entropy + topic entropy + style coverage)
    topic = (brain or {}).get("topic_ema") or {}
    topic_counter = Counter({k: float(v) for k, v in topic.items() if isinstance(v, (int, float))})
    topic_ent = norm_entropy(topic_counter) if topic_counter else 0.0

    if archive_path.exists():
        cats = category_counter_from_archive(archive_path)
        used = style_usage_from_archive(archive_path)
    else:
        cats = Counter(str(it.get("category") or "") for it in memory if isinstance(it, dict))
        used = set()
        for it in memory:
            if not isinstance(it, dict):
                continue
            if str(it.get("action") or "") == "comment":
                s = str(it.get("used_strategy") or "")
                t = str(it.get("used_tone") or "")
                l = str(it.get("used_length") or "")
                tid = str(it.get("template_id") or "")
                if s:
                    used.add(f"strategy:{s}")
                if t:
                    used.add(f"tone:{t}")
                if l:
                    used.add(f"comment_length:{l}")
                if tid:
                    used.add(f"template:{tid}")
            elif str(it.get("action") or "") == "post":
                sty = str(it.get("used_style") or "")
                if sty:
                    used.add(f"post_styles:{sty}")

    total_opts = count_policy_options(policy if isinstance(policy, dict) else {})
    style_cov = (len(used) / total_opts) if total_opts else 0.0

    cat_ent = norm_entropy(cats)
    act_ent = norm_entropy(acts_life)
    diversity = clamp(100 * (0.30 * cat_ent + 0.20 * act_ent + 0.30 * topic_ent + 0.20 * style_cov))

    # REL: reliability (vote duplicate avoidance + recent eval success)
    unique_vote_posts = 0
    vp = (state or {}).get("voted_posts")
    if isinstance(vp, dict):
        unique_vote_posts = len(vp)

    dupe_rate = max(0.0, 1.0 - (unique_vote_posts / votes_life)) if votes_life else 0.0
    eval_ok_recent = compute_recent_eval_success(memory)
    eval_ok_score = eval_ok_recent if eval_ok_recent is not None else 0.5
    reliability = clamp(100 * (0.60 * (1.0 - dupe_rate) + 0.40 * eval_ok_score))

    # LRN: learning progress (lifetime evaluated fraction + policy drift)
    eval_frac_life = (evaluated_count / total_contrib_life) if total_contrib_life else 0.0
    drift_frac = policy_drift_fraction(policy if isinstance(policy, dict) else {})
    learning = clamp(100 * (0.70 * eval_frac_life + 0.30 * drift_frac))

    # RISK: exposure to injection-category (penalty-like) over contributions
    inj = int(cats.get("injection", 0))
    inj_ratio = (inj / sum(cats.values())) if sum(cats.values()) else 0.0
    risk = clamp(100 * (inj_ratio / 0.10))  # 10% injection exposure -> 100

    # Composite Growth Index (0..100)
    GI = clamp(
        0.23 * experience
        + 0.22 * judgment
        + 0.20 * engagement
        + 0.18 * diversity
        + 0.12 * reliability
        + 0.05 * learning
        - 0.10 * risk
    )
    IQ = 60 + 0.8 * GI

    stage = "Bootstrapping" if total_actions_life < 200 else "Learning" if total_actions_life < 1000 else "Mature"

    # Today's counters
    by_day = ((semantic or {}).get("by_day") or {}).get(date_kst, {}) if isinstance(semantic, dict) else {}
    today_ticks = int(by_day.get("ticks", 0) or 0)
    today_posts = int(by_day.get("post", 0) or 0)
    today_comments_root = int(by_day.get("comment", 0) or 0)
    today_replies_other = int(by_day.get("reply_other", 0) or 0)
    today_replies_own = int(by_day.get("reply_own", 0) or 0)
    today_replies = int(today_replies_other + today_replies_own)
    today_comments = int(today_comments_root + today_replies)
    today_votes = int((by_day.get("vote:up", 0) or 0) + (by_day.get("vote:down", 0) or 0))
    today_eval = int(by_day.get("eval", 0) or 0)

    rr = compute_recent_reward_stats(memory)

    def safe_len_dict(d: Any) -> int:
        return len(d) if isinstance(d, dict) else 0

    def safe_len_list(x: Any) -> int:
        return len(x) if isinstance(x, list) else 0

    snap: Dict[str, Any] = {
        "Snapshot (KST)": now,
        "Date (KST)": date_kst,
        "Stage": stage,

        "Mersoom IQ (internal)": round(IQ, 1),
        "Growth Index (0-100)": round(GI, 1),

        "EXP (Experience)": round(experience, 1),
        "JDG (Judgment)": round(judgment, 1),
        "DEP (Depth)": round(engagement, 1),
        "DIV (Diversity)": round(diversity, 1),
        "REL (Reliability)": round(reliability, 1),
        "LRN (Learning)": round(learning, 1),
        "RISK (Exposure)": round(risk, 1),

        # Lifetime totals
        "Total Actions (lifetime)": total_actions_life,
        "Votes (lifetime)": votes_life,
        "Posts (lifetime)": posts_life,
        "Comments (lifetime)": comments_life,
        "Comments root (lifetime)": comments_root_life,
        "Replies (lifetime)": int(replies_other_life + replies_own_life),
        "Replies to others (lifetime)": replies_other_life,
        "Replies to own (lifetime)": replies_own_life,
        "Evaluated contrib (lifetime)": evaluated_count,
        "Avg reward (lifetime)": round(avg_reward, 4) if avg_reward is not None else None,

        # Recent window
        "Window size (memory.json)": int(len(memory)),
        "Votes (recent)": int(acts_recent.get("vote:up", 0) + acts_recent.get("vote:down", 0)),
        "Posts (recent)": int(acts_recent.get("post", 0)),
        "Comments (recent)": int(comments_root_recent),
        "Replies (recent)": int(replies_recent),
        "Replies to others (recent)": int(replies_other_recent),
        "Replies to own (recent)": int(replies_own_recent),
        "Reward mean (recent, contrib)": round(rr["mean"], 4) if rr["mean"] is not None else None,
        "Reward pos-rate (recent, contrib)": round(rr["pos_rate"], 3) if rr["pos_rate"] is not None else None,
        "Eval success (recent)": round(eval_ok_recent, 3) if eval_ok_recent is not None else None,

        "Style coverage": round(style_cov, 3),
        "Category entropy (norm)": round(cat_ent, 3),
        "Action entropy (norm)": round(act_ent, 3),
        "Topic entropy (norm)": round(topic_ent, 3),

        "Policy epsilon": (policy or {}).get("epsilon") if isinstance(policy, dict) else None,
        "Policy lr": (policy or {}).get("lr") if isinstance(policy, dict) else None,

        "Seen Posts": safe_len_list((state or {}).get("seen_post_ids")),
        "Voted Posts": safe_len_dict((state or {}).get("voted_posts")),
        "Replied Posts": safe_len_dict((state or {}).get("replied_ts")),
        "Commented Posts": safe_len_dict((state or {}).get("commented_ts")),

        # Today
        "Today ticks": today_ticks,
        "Today votes": today_votes,
        "Today posts": today_posts,
        "Today comments": today_comments,
        "Today root comments": today_comments_root,
        "Today replies": today_replies,
        "Today replies to others": today_replies_other,
        "Today replies to own": today_replies_own,
        "Today eval": today_eval,

        "Notes": "",
    }
    return snap


# ---- Excel helpers ----
def autosize(ws: Worksheet, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 500) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), max_width)


def _header_style() -> Tuple[PatternFill, Font, Alignment]:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return header_fill, header_font, header_align


def _apply_header_row_style(ws: Worksheet, header_row: int, headers: List[str]) -> None:
    fill, font, align = _header_style()
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[header_row].height = 28


def _set_number_formats_for_row(ws: Worksheet, headers: List[str], row_idx: int) -> None:
    # timestamp
    if "Snapshot (KST)" in headers:
        ws.cell(row=row_idx, column=headers.index("Snapshot (KST)") + 1).number_format = "yyyy-mm-dd hh:mm:ss"

    one_decimal = {
        "Mersoom IQ (internal)", "Growth Index (0-100)",
        "EXP (Experience)", "JDG (Judgment)", "DEP (Depth)", "DIV (Diversity)",
        "REL (Reliability)", "LRN (Learning)", "RISK (Exposure)",
    }
    three_decimal = {
        "Style coverage", "Category entropy (norm)", "Action entropy (norm)", "Topic entropy (norm)",
        "Reward pos-rate (recent, contrib)", "Eval success (recent)",
    }
    four_decimal = {"Avg reward (lifetime)", "Reward mean (recent, contrib)"}

    for h in one_decimal:
        if h in headers:
            ws.cell(row=row_idx, column=headers.index(h) + 1).number_format = "0.0"
    for h in three_decimal:
        if h in headers:
            ws.cell(row=row_idx, column=headers.index(h) + 1).number_format = "0.000"
    for h in four_decimal:
        if h in headers:
            ws.cell(row=row_idx, column=headers.index(h) + 1).number_format = "0.0000"


def _ensure_metrics_table(ws: Worksheet, headers: List[str]) -> None:
    """Ensure a single table exists and covers all rows/cols."""
    last_row = ws.max_row
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{last_row}"

    # remove extra tables if any
    for tname in list(ws.tables.keys()):
        try:
            del ws.tables[tname]
        except Exception:
            pass

    table = Table(displayName="MetricsTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)


def _ensure_dashboard(wb: Workbook, snapshot: Dict[str, Any]) -> Worksheet:
    if "Dashboard" in wb.sheetnames:
        dash = wb["Dashboard"]
    else:
        dash = wb.create_sheet("Dashboard")

    dash["A1"] = "Latest snapshot"
    dash["A1"].font = Font(bold=True, size=14)

    key_fields = [
        "Snapshot (KST)", "Stage",
        "Mersoom IQ (internal)", "Growth Index (0-100)",
        "EXP (Experience)", "JDG (Judgment)", "DEP (Depth)", "DIV (Diversity)",
        "REL (Reliability)", "LRN (Learning)", "RISK (Exposure)",

        "Total Actions (lifetime)", "Votes (lifetime)", "Posts (lifetime)", "Comments (lifetime)", "Comments root (lifetime)", "Replies (lifetime)", "Replies to others (lifetime)", "Replies to own (lifetime)",
        "Evaluated contrib (lifetime)", "Avg reward (lifetime)",

        "Window size (memory.json)", "Votes (recent)", "Posts (recent)", "Comments (recent)", "Replies (recent)", "Replies to others (recent)", "Replies to own (recent)",
        "Reward mean (recent, contrib)", "Reward pos-rate (recent, contrib)", "Eval success (recent)",

        "Style coverage",
        "Today ticks", "Today votes", "Today posts", "Today comments", "Today root comments", "Today replies", "Today replies to others", "Today replies to own", "Today eval",
    ]

    # clear old content region (lightly)
    for r in range(3, 3 + 120):
        dash[f"A{r}"].value = None
        dash[f"B{r}"].value = None

    for i, k in enumerate(key_fields, start=3):
        dash[f"A{i}"] = k
        dash[f"A{i}"].font = Font(bold=True)
        dash[f"B{i}"] = snapshot.get(k)

    dash.column_dimensions["A"].width = 34
    dash.column_dimensions["B"].width = 28

    # timestamp format
    dash["B3"].number_format = "yyyy-mm-dd hh:mm:ss"

    # Radar source + chart area
    radar_dims = ["EXP (Experience)", "JDG (Judgment)", "DEP (Depth)", "DIV (Diversity)", "REL (Reliability)", "LRN (Learning)"]
    dash["D2"] = "Dimension"
    dash["E2"] = "Score"
    dash["D2"].font = dash["E2"].font = Font(bold=True)

    for r, dim in enumerate(radar_dims, start=3):
        dash[f"D{r}"] = dim.split(" ")[0]  # EXP/JDG/...
        dash[f"E{r}"] = snapshot.get(dim)
        dash[f"E{r}"].number_format = "0.0"

    # remove existing charts on dashboard
    try:
        dash._charts = []
    except Exception:
        pass

    radar = RadarChart()
    radar.title = "Multi-dim benchmark (0-100)"
    data = Reference(dash, min_col=5, min_row=2, max_row=2 + len(radar_dims))
    cats = Reference(dash, min_col=4, min_row=3, max_row=2 + len(radar_dims))
    radar.add_data(data, titles_from_data=True)
    radar.set_categories(cats)
    radar.y_axis.scaling.min = 0
    radar.y_axis.scaling.max = 100
    radar.height = 14
    radar.width = 18
    dash.add_chart(radar, "D8")

    return dash


def _ensure_definitions(wb: Workbook) -> Worksheet:
    if "Definitions" in wb.sheetnames:
        defs = wb["Definitions"]
        # clear old
        for r in range(1, min(defs.max_row + 1, 300)):
            defs[f"A{r}"].value = None
    else:
        defs = wb.create_sheet("Definitions")

    defs["A1"] = "Definitions"
    defs["A1"].font = Font(bold=True, size=14)

    lines = [
        "IQ는 사람 IQ가 아니라, 에이전트 상태를 요약한 내부 지표입니다.",
        "Growth Index(0~100)는 아래 구성요소(0~100)를 합성한 값이며, IQ = 60 + 0.8 * GrowthIndex 입니다.",
        "",
        "[데이터 범위]",
        "- lifetime(누적) 지표는 주로 state/semantic을 사용합니다.",
        "  * state.total_actions / state.total_reward / state.evaluated_count",
        "  * semantic.by_day의 vote:up/down, post, comment, eval, ticks",
        "- memory.json은 최근 윈도우만 남습니다(rolling).",
        "- memory_archive.jsonl이 있으면 post/comment에 한해 스타일/카테고리 누적 집계에 사용합니다.",
        "",
        "[구성요소]",
        "- EXP: lifetime 총 actions가 쌓일수록 증가 (log 스케일)",
        "- JDG: lifetime 평균 reward(state.total_reward/state.evaluated_count)를 0~100으로 변환(없으면 0)",
        "- DEP: lifetime에서 post+comment 비중(깊은 상호작용)",
        "- DIV: (카테고리/행동/토픽/스타일 커버리지) 혼합",
        "- REL: (누적) 중복 투표 방지 + (최근 윈도우) 평가 성공률(없으면 중립값 0.5)",
        "- LRN: (누적) 평가 누적률 + 정책 가중치 드리프트(학습 시작 신호)",
        "- RISK: injection 카테고리 비중(과도하면 패널티)",
        "",
        "운영 초반에는 evaluated_count가 0일 수 있고, 이 경우 JDG/LRN이 낮게 나오는 게 정상입니다.",
    ]
    for i, line in enumerate(lines, start=3):
        defs.cell(row=i, column=1, value=line)
    defs.column_dimensions["A"].width = 120
    return defs


def rebuild_charts_sheet(wb: Workbook, ws_data: Worksheet, headers: List[str]) -> None:
    """Recreate the Charts sheet so chart ranges always include the latest appended rows."""
    if "Charts" in wb.sheetnames:
        wb.remove(wb["Charts"])
    charts = wb.create_sheet("Charts")
    charts["A1"] = "Trends"
    charts["A1"].font = Font(bold=True, size=14)

    # pick time column
    time_header = "Snapshot (KST)"
    time_col = headers.index(time_header) + 1 if time_header in headers else 1

    def add_chart(title: str, y_header: str, anchor: str) -> None:
        if y_header not in headers:
            return
        y_col = headers.index(y_header) + 1
        c = LineChart()
        c.title = title
        data_ref = Reference(ws_data, min_col=y_col, min_row=1, max_row=ws_data.max_row)
        cats_ref = Reference(ws_data, min_col=time_col, min_row=2, max_row=ws_data.max_row)
        c.add_data(data_ref, titles_from_data=True)
        c.set_categories(cats_ref)
        c.height = 10
        c.width = 26
        charts.add_chart(c, anchor)

    # Primary
    add_chart("Mersoom IQ", "Mersoom IQ (internal)", "A3")
    add_chart("Growth Index", "Growth Index (0-100)", "A18")
    add_chart("Avg reward (lifetime)", "Avg reward (lifetime)", "A33")

    # Optional counts
    add_chart("Total Actions (lifetime)", "Total Actions (lifetime)", "AA3")
    add_chart("Votes (lifetime)", "Votes (lifetime)", "AA18")
    add_chart("Posts (lifetime)", "Posts (lifetime)", "AA33")


def init_workbook(xlsx_path: Path, snapshot: Dict[str, Any]) -> None:
    headers = list(snapshot.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics Log"

    ws.append(headers)
    ws.append([snapshot.get(h) for h in headers])

    _apply_header_row_style(ws, 1, headers)
    ws.freeze_panes = "A2"

    _set_number_formats_for_row(ws, headers, 2)
    _ensure_metrics_table(ws, headers)
    autosize(ws)

    _ensure_dashboard(wb, snapshot)
    _ensure_definitions(wb)
    rebuild_charts_sheet(wb, ws, headers)

    wb.save(xlsx_path)


def _migrate_headers(ws: Worksheet, existing_headers: List[Any], desired_headers: List[str]) -> List[str]:
    """Return merged headers and update sheet if new headers are needed."""
    existing = [str(x) if x is not None else "" for x in existing_headers]
    existing = [h for h in existing if h]

    merged = list(existing)
    for h in desired_headers:
        if h not in merged:
            merged.append(h)

    if merged == existing:
        return merged

    # write new headers
    for col, h in enumerate(merged, start=1):
        ws.cell(row=1, column=col).value = h

    _apply_header_row_style(ws, 1, merged)
    return merged


def append_snapshot(xlsx_path: Path, snapshot: Dict[str, Any]) -> None:
    desired_headers = list(snapshot.keys())

    if not xlsx_path.exists():
        init_workbook(xlsx_path, snapshot)
        return

    wb = load_workbook(xlsx_path)

    # Ensure Metrics Log exists
    if "Metrics Log" not in wb.sheetnames:
        # create new minimal structure inside the existing workbook
        ws = wb.create_sheet("Metrics Log", 0)
        ws.append(desired_headers)
        _apply_header_row_style(ws, 1, desired_headers)
        ws.freeze_panes = "A2"
    ws = wb["Metrics Log"]

    existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    merged_headers = _migrate_headers(ws, existing_headers, desired_headers)

    # append values in merged header order
    row_values = [snapshot.get(h) for h in merged_headers]
    ws.append(row_values)
    new_row = ws.max_row

    _set_number_formats_for_row(ws, merged_headers, new_row)

    _ensure_metrics_table(ws, merged_headers)
    autosize(ws)

    # Update / (re)create Dashboard + Definitions
    _ensure_dashboard(wb, snapshot)
    _ensure_definitions(wb)

    rebuild_charts_sheet(wb, ws, merged_headers)
    wb.save(xlsx_path)


def main() -> None:
    default_dir = r"D:\강준규\업무정리\##개인작업"

    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", type=str, default=default_dir, help="Folder that contains mersoom_*.json")
    ap.add_argument("--out", type=str, default="", help="Output xlsx path (default: <dir>\\mersoom_agent_iq_scorecard.xlsx or env MERSOOM_SCORECARD)")
    args = ap.parse_args()

    files_dir = Path(args.dir)
    if not files_dir.exists():
        raise SystemExit(f"[ERROR] Directory not found: {files_dir}")

    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = (files_dir / out_path).resolve()
    else:
        out_path = path_from_env(files_dir, "MERSOOM_SCORECARD", "mersoom_agent_iq_scorecard.xlsx")

    snapshot = build_snapshot(files_dir)
    append_snapshot(out_path, snapshot)

    print("✅ Snapshot appended.")
    print(f"- File: {out_path}")
    print(f"- IQ: {snapshot.get('Mersoom IQ (internal)')}, GrowthIndex: {snapshot.get('Growth Index (0-100)')}")
    print(f"- Lifetime actions: {snapshot.get('Total Actions (lifetime)')} | votes={snapshot.get('Votes (lifetime)')} posts={snapshot.get('Posts (lifetime)')} comments={snapshot.get('Comments (lifetime)')}")
    print(f"- Window: {snapshot.get('Window size (memory.json)')} | recent votes={snapshot.get('Votes (recent)')} posts={snapshot.get('Posts (recent)')} comments={snapshot.get('Comments (recent)')} replies={snapshot.get('Replies (recent)')}")


if __name__ == "__main__":
    main()
